#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Anthropic 提示工程教程答案键 - HTML转Excel转换器
将HTML格式的教程答案键转换为Excel文件
"""

import pandas as pd
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_excel_from_data():
    """直接从数据创建Excel文件"""
    
    # 准备数据
    data = [
        ["", "欢迎来到提示工程交互式教程", "", ""],
        ["", "注意：这是教程的非交互式答案键！您将无法在此表格中与Claude进行交互。", "", ""],
        ["", "", "", ""],
        ["", "课程介绍和目标", "", ""],
        ["", "本课程旨在为您提供全面的逐步指导，教您如何在Claude中设计最优的提示词。", "", ""],
        ["", "完成本课程后，您将能够：", "", ""],
        ["", "✓ 掌握优秀提示词的基本结构", "", ""],
        ["", "✓ 识别常见失败模式并学习解决这些问题的\"80/20\"技巧", "", ""],
        ["", "✓ 了解Claude的优势和局限性", "", ""],
        ["", "✓ 从零开始为常见用例构建强大的提示词", "", ""],
        ["", "", "", ""],
        ["", "课程结构和内容", "", ""],
        ["", "本课程的结构设计让您有很多机会亲自练习编写和调试提示词。", "", ""],
        ["", "课程分为9个章节和配套练习，以及一个包含更高级方法的附录。", "", ""],
        ["", "建议您按章节顺序学习课程。", "", ""],
        ["", "", "", ""],
        ["", "初级", "", ""],
        ["1", "第1章：基本提示词结构", "章节", ""],
        ["", "课程", "课程内容", "学习提示词的基本构成要素"],
        ["", "练习", "实践练习", "动手练习基本提示词结构"],
        ["2", "第2章：清晰直接的表达", "章节", ""],
        ["", "课程", "课程内容", "学习如何使提示词更加清晰明确"],
        ["", "练习", "实践练习", "练习编写清晰直接的提示词"],
        ["3", "第3章：角色分配（角色提示）", "章节", ""],
        ["", "课程", "课程内容", "学习如何为Claude分配特定角色"],
        ["", "练习", "实践练习", "练习角色分配技巧"],
        ["", "", "", ""],
        ["", "中级", "", ""],
        ["4", "第4章：将数据与指令分离", "章节", ""],
        ["", "课程", "课程内容", "学习如何有效分离数据和指令"],
        ["", "练习", "实践练习", "练习数据与指令分离技术"],
        ["5", "第5章：格式化输出和为Claude代言", "章节", ""],
        ["", "课程", "课程内容", "学习如何控制输出格式"],
        ["", "练习", "实践练习", "练习格式化输出技巧"],
        ["6", "第6章：预认知（逐步思考）", "章节", ""],
        ["", "课程", "课程内容", "学习如何引导Claude逐步思考"],
        ["", "练习", "实践练习", "练习逐步思考技术"],
        ["7", "第7章：使用示例（小样本提示）", "章节", ""],
        ["", "课程", "课程内容", "学习如何使用示例提升效果"],
        ["", "练习", "实践练习", "练习小样本提示技术"],
        ["", "", "", ""],
        ["", "高级", "", ""],
        ["8", "第8章：避免幻觉", "章节", ""],
        ["", "课程", "课程内容", "学习如何减少AI幻觉现象"],
        ["", "练习", "实践练习", "练习避免幻觉的技巧"],
        ["9", "第9章：构建复杂提示词（行业用例）", "章节", ""],
        ["", "从零开始构建复杂提示词 - 聊天机器人", "课程内容", "实际案例：聊天机器人构建"],
        ["", "法律服务的复杂提示词", "课程内容", "实际案例：法律行业应用"],
        ["", "练习：金融服务的复杂提示词", "实践练习", "金融行业应用练习"],
        ["", "练习：编程的复杂提示词", "实践练习", "编程领域应用练习"],
        ["", "恭喜您完成课程 & 下一步", "总结", "课程总结和后续建议"],
        ["", "", "", ""],
        ["", "附录：超越标准提示", "", ""],
        ["A1", "链式提示", "高级技术", "学习如何链接多个提示"],
        ["A2", "函数调用", "高级技术", "学习如何使用函数调用功能"],
        ["A3", "搜索和检索", "高级技术", "学习搜索和检索增强技术"],
        ["", "", "", ""],
        ["", "→ 第1章：基本提示词结构", "", ""]
    ]
    
    # 创建DataFrame
    df = pd.DataFrame(data, columns=["序号", "章节内容", "类型", "说明"])
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anthropic提示工程教程答案键"
    
    # 添加标题
    ws.merge_cells('A1:D1')
    ws['A1'] = '[答案键] Anthropic 提示工程交互式教程 [公开访问] - 中文版'
    title_font = Font(name='微软雅黑', size=16, bold=True, color='1F497D')
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # 添加警告信息
    ws.merge_cells('A2:D2')
    ws['A2'] = '注意：这是教程的非交互式答案键！您将无法在此表格中与Claude进行交互。'
    warning_font = Font(name='微软雅黑', size=10, color='D68910')
    warning_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    ws['A2'].font = warning_font
    ws['A2'].fill = warning_fill
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25
    
    # 添加表头（从第4行开始）
    headers = ["序号", "章节内容", "类型", "说明"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(name='微软雅黑', size=12, bold=True)
        cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 添加数据（从第5行开始）
    for row_idx, row_data in enumerate(data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # 根据内容类型设置样式
            if col_idx == 2 and value:  # 章节内容列
                if "第" in value and "章" in value:
                    cell.font = Font(name='微软雅黑', size=11, bold=True)
                    cell.fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
                elif value in ["初级", "中级", "高级", "附录：超越标准提示"]:
                    cell.font = Font(name='微软雅黑', size=12, bold=True, color='1F497D')
                    cell.fill = PatternFill(start_color='D1E7DD', end_color='D1E7DD', fill_type='solid')
                    ws.merge_cells(f'A{row_idx}:D{row_idx}')
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    
    # 添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
    
    # 保存文件
    filename = "Anthropic提示工程教程答案键_中文版.xlsx"
    wb.save(filename)
    print(f"✅ Excel文件已创建：{filename}")
    return filename

def main():
    """主函数"""
    print("🚀 开始创建Anthropic提示工程教程答案键的Excel文件...")
    
    try:
        filename = create_excel_from_data()
        print(f"\n✅ 成功创建Excel文件：{filename}")
        print(f"📁 文件位置：{os.path.abspath(filename)}")
        print("\n📝 使用说明：")
        print("1. 双击打开Excel文件")
        print("2. 可以直接编辑和打印")
        print("3. 可以导入到Google Sheets等其他电子表格应用")
        
    except Exception as e:
        print(f"❌ 创建Excel文件时出错：{e}")
        print("\n💡 如果缺少依赖库，请运行以下命令安装：")
        print("pip install pandas openpyxl beautifulsoup4")

if __name__ == "__main__":
    main() 